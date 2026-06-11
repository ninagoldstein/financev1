import pytest
from openpyxl import Workbook, load_workbook

from secpull.income_model import (
    MarginAssumptions,
    avg_margin,
    build_margin_assumptions,
    historical_metrics,
    implied_shares_m,
    project_income_statement,
)
from secpull.models import FinancialFact


def _fact(metric, value, fy, end, filed="2024-01-01", period="FY", form="10-K"):
    unit = "USD/shares" if metric == "eps_diluted" else "USD"
    return FinancialFact(
        cik="0001397187", metric=metric, tag_used="x",
        value=value, unit=unit, fiscal_year=fy, fiscal_period=period,
        form=form, end_date=end, filed_date=filed,
    )


FACTS = [
    _fact("revenue",          8.111e9, 2022, "2023-01-29"),
    _fact("revenue",          9.619e9, 2023, "2024-01-28"),
    _fact("gross_profit",     4.700e9, 2022, "2023-01-29"),
    _fact("gross_profit",     5.600e9, 2023, "2024-01-28"),
    _fact("operating_income", 1.600e9, 2022, "2023-01-29"),
    _fact("operating_income", 1.900e9, 2023, "2024-01-28"),
    _fact("net_income",       1.200e9, 2022, "2023-01-29"),
    _fact("net_income",       1.550e9, 2023, "2024-01-28"),
    _fact("eps_diluted",      9.64,    2022, "2023-01-29"),
    _fact("eps_diluted",      12.20,   2023, "2024-01-28"),
]

REV_SERIES = [(2022, 8.111e9), (2023, 9.619e9)]


# ── historical_metrics ────────────────────────────────────────────────────────

def test_historical_metrics_fy_only():
    quarterly = _fact("gross_profit", 1e9, 2023, "2023-04-30", period="Q1", form="10-Q")
    m = historical_metrics(FACTS + [quarterly], "gross_profit")
    assert 2023 in m
    assert m[2023] == 5.600e9  # FY value, not quarterly

def test_historical_metrics_latest_filed_wins():
    restatement = _fact("gross_profit", 9.999e9, 2022, "2023-01-29", filed="2025-03-01")
    m = historical_metrics(FACTS + [restatement], "gross_profit")
    assert m[2022] == 9.999e9

def test_historical_metrics_empty_when_absent():
    m = historical_metrics(FACTS, "total_assets")
    assert m == {}


# ── avg_margin ────────────────────────────────────────────────────────────────

def test_avg_margin_basic():
    assert abs(avg_margin([0.10, 0.20, 0.30]) - 0.20) < 1e-9

def test_avg_margin_skips_none():
    assert abs(avg_margin([None, 0.10, None, 0.30]) - 0.20) < 1e-9

def test_avg_margin_all_none_returns_none():
    assert avg_margin([None, None]) is None

def test_avg_margin_empty_returns_none():
    assert avg_margin([]) is None

def test_avg_margin_uses_last_five():
    values = [0.50, 0.10, 0.20, 0.30, 0.40, 0.05]  # 6 values
    result = avg_margin(values, n=5)
    expected = (0.10 + 0.20 + 0.30 + 0.40 + 0.05) / 5
    assert abs(result - expected) < 1e-9


# ── implied_shares_m ──────────────────────────────────────────────────────────

def test_implied_shares_m_basic():
    # net_income=$1.55B, eps=$12.20 → 127.05M shares
    shares = implied_shares_m(1.55e9, 12.20)
    assert abs(shares - (1.55e9 / 12.20 / 1e6)) < 1e-6

def test_implied_shares_m_zero_eps_returns_none():
    assert implied_shares_m(1e9, 0.0) is None


# ── build_margin_assumptions ──────────────────────────────────────────────────

def test_build_margin_assumptions_full_data():
    a = build_margin_assumptions(FACTS, REV_SERIES)
    assert isinstance(a, MarginAssumptions)
    assert a.gross_margin is not None
    assert a.operating_margin is not None
    assert a.net_margin is not None
    assert a.diluted_shares_m is not None
    expected_gm = ((4.7e9 / 8.111e9) + (5.6e9 / 9.619e9)) / 2
    assert abs(a.gross_margin - expected_gm) < 1e-4

def test_build_margin_assumptions_missing_gross_profit():
    facts_no_gp = [f for f in FACTS if f.metric != "gross_profit"]
    a = build_margin_assumptions(facts_no_gp, REV_SERIES)
    assert a.gross_margin is None
    assert a.net_margin is not None  # other metrics unaffected

def test_build_margin_assumptions_uses_most_recent_shares():
    a = build_margin_assumptions(FACTS, REV_SERIES)
    # Most recent year (2023): 1.55e9 / 12.20 / 1e6
    expected = 1.55e9 / 12.20 / 1e6
    assert abs(a.diluted_shares_m - expected) < 1e-4


# ── project_income_statement ──────────────────────────────────────────────────

_MA = MarginAssumptions(
    gross_margin=0.50, operating_margin=0.20, net_margin=0.10,
    diluted_shares_m=100.0, n_years_used=2,
)
_REV_PROJ = {"base": [(2024, 1_000_000_000.0)]}  # $1B, one year


def test_project_income_statement_values_correct():
    proj = project_income_statement(_REV_PROJ, _MA)
    base = proj["base"]
    assert abs(base["gross_profit"][0][1] - 5e8) < 1.0
    assert abs(base["net_income"][0][1] - 1e8) < 1.0
    # EPS = 1e8 raw $ / (100M shares) = $1.00
    assert abs(base["eps_diluted"][0][1] - 1.00) < 0.001

def test_project_income_statement_all_scenarios_present():
    proj = project_income_statement(
        {"bear": [(2024, 9e8)], "base": [(2024, 1e9)], "bull": [(2024, 1.1e9)]}, _MA
    )
    assert set(proj.keys()) == {"bear", "base", "bull"}

def test_project_income_statement_none_margin_yields_none():
    ma = MarginAssumptions(
        gross_margin=None, operating_margin=0.20, net_margin=0.10,
        diluted_shares_m=100.0, n_years_used=2,
    )
    proj = project_income_statement(_REV_PROJ, ma)
    assert proj["base"]["gross_profit"][0][1] is None
    assert proj["base"]["net_income"][0][1] is not None

def test_project_income_statement_none_shares_gives_none_eps():
    ma = MarginAssumptions(
        gross_margin=0.50, operating_margin=0.20, net_margin=0.10,
        diluted_shares_m=None, n_years_used=2,
    )
    proj = project_income_statement(_REV_PROJ, ma)
    assert proj["base"]["eps_diluted"][0][1] is None


# ── add_income_sheet ──────────────────────────────────────────────────────────

def _make_rev_assump():
    from secpull.model import build_assumptions
    return build_assumptions(0.10, offset=0.05, n_years_used=3)


def test_add_income_sheet_creates_sheet(tmp_path):
    from secpull.income_model_export import add_income_sheet
    wb = Workbook()
    metric_series = {
        m: historical_metrics(FACTS, m)
        for m in ("gross_profit", "operating_income", "net_income", "eps_diluted")
    }
    ma = build_margin_assumptions(FACTS, REV_SERIES)
    proj_years = [2024, 2025, 2026]
    add_income_sheet(wb, "LULU", REV_SERIES, metric_series, ma, _make_rev_assump(), proj_years)
    out = tmp_path / "lulu.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    assert "Income Statement" in wb2.sheetnames


def test_add_income_sheet_projection_formulas(tmp_path):
    from secpull.income_model_export import add_income_sheet
    wb = Workbook()
    metric_series = {
        m: historical_metrics(FACTS, m)
        for m in ("gross_profit", "operating_income", "net_income", "eps_diluted")
    }
    ma = build_margin_assumptions(FACTS, REV_SERIES)
    add_income_sheet(wb, "LULU", REV_SERIES, metric_series, ma, _make_rev_assump(), [2024, 2025, 2026])
    out = tmp_path / "lulu.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["Income Statement"]
    # Bear revenue year 1 (row 30, col B) must be a formula
    assert str(ws["B30"].value).startswith("=")
    # Bear gross profit year 1 (row 31, col B) must be a formula
    assert str(ws["B31"].value).startswith("=")


def test_add_income_sheet_assumptions_numeric(tmp_path):
    from secpull.income_model_export import add_income_sheet
    wb = Workbook()
    metric_series = {
        m: historical_metrics(FACTS, m)
        for m in ("gross_profit", "operating_income", "net_income", "eps_diluted")
    }
    ma = build_margin_assumptions(FACTS, REV_SERIES)
    add_income_sheet(wb, "LULU", REV_SERIES, metric_series, ma, _make_rev_assump(), [2024, 2025, 2026])
    out = tmp_path / "lulu.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["Income Statement"]
    assert isinstance(ws["B20"].value, float)  # gross margin assumption
    assert isinstance(ws["B17"].value, float)  # bear growth


def test_add_income_sheet_na_when_metric_absent(tmp_path):
    from secpull.income_model_export import add_income_sheet
    wb = Workbook()
    metric_series = {
        m: historical_metrics(FACTS, m)
        for m in ("gross_profit", "operating_income", "net_income", "eps_diluted")
    }
    # Build assumptions with gross_margin forced to None
    ma = MarginAssumptions(
        gross_margin=None, operating_margin=0.20, net_margin=0.15,
        diluted_shares_m=127.0, n_years_used=2,
    )
    add_income_sheet(wb, "LULU", REV_SERIES, metric_series, ma, _make_rev_assump(), [2024, 2025, 2026])
    out = tmp_path / "lulu.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    ws = wb2["Income Statement"]
    assert ws["B20"].value == "N/A"   # gross margin assumption cell
    assert ws["B31"].value == "N/A"   # bear gross profit projection year 1
