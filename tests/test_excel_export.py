"""Tests for secpull/excel_export.py."""
import json
import pathlib
import tempfile

import openpyxl
import pytest

from secpull import config
from secpull.assumptions import build_assumptions_from_profile
from secpull.dcf import DCFInputs, build_dcf
from secpull.derived import compute_derived_metrics
from secpull.excel_export import export_workbook
from secpull.extract import extract_metrics
from secpull.forecast import build_forecast
from secpull.models import FinancialFact
from secpull.profile import build_profile
from secpull.quality import COMPLETE
from secpull.statements import build_statements

# ── Expected sheet names ──────────────────────────────────────────────────────

EXPECTED_SHEETS = [
    "Summary",
    "Historical Financials",
    "Forecast",
    "DCF",
    "DCF Multiple",
    "Sensitivity - GGM",
    "Sensitivity - Exit",
    "Assumption Audit",
]

# ── Shared pipeline fixture ───────────────────────────────────────────────────

_CIK   = "0001397187"
_TICKER = "LULU"


def _f(metric, fy, value, quality=COMPLETE, filed="2026-02-01"):
    return FinancialFact(
        cik="0001111111", metric=metric, tag_used="Tag",
        value=value, unit="USD",
        fiscal_year=fy, fiscal_period="FY",
        form="10-K", end_date=f"{fy + 1}-01-31", filed_date=filed,
        coverage_quality=quality,
    )


@pytest.fixture(scope="module")
def lulu_pipeline():
    """Full pipeline from raw JSON → DCFResult for LULU."""
    with open(config.DATA_DIR / "raw" / f"{_CIK}.json") as f:
        payload = json.load(f)
    facts   = extract_metrics(_CIK, payload)
    derived = compute_derived_metrics(_CIK, facts)
    stmts   = build_statements(_CIK, _TICKER, facts, derived, max_years=5)
    profile = build_profile(stmts, "lululemon athletica")
    asmp    = build_assumptions_from_profile(profile)
    proj    = build_forecast(asmp, profile)
    inputs  = DCFInputs(
        wacc=0.10,
        terminal_growth_rate=0.025,
        exit_ebitda_multiple=10.0,
        net_debt=-1_500e6,        # -$1.5B net cash in raw USD
        diluted_shares=125_000_000,  # actual share count
    )
    dcf = build_dcf(proj, inputs)
    return stmts, profile, asmp, proj, dcf


@pytest.fixture(scope="module")
def lulu_workbook(lulu_pipeline, tmp_path_factory):
    """Generate the workbook once and return the loaded openpyxl Workbook."""
    stmts, profile, asmp, proj, dcf = lulu_pipeline
    out = tmp_path_factory.mktemp("xl") / f"{_TICKER}_test.xlsx"
    path = export_workbook(stmts, profile, asmp, proj, dcf, output_path=out)
    wb = openpyxl.load_workbook(path)
    return wb, path


# ── File creation ─────────────────────────────────────────────────────────────

def test_workbook_file_created(lulu_workbook):
    wb, path = lulu_workbook
    assert path.exists()
    assert path.suffix == ".xlsx"
    assert path.stat().st_size > 0


def test_export_returns_path(lulu_pipeline, tmp_path):
    stmts, profile, asmp, proj, dcf = lulu_pipeline
    out = tmp_path / "test.xlsx"
    result = export_workbook(stmts, profile, asmp, proj, dcf, output_path=out)
    assert isinstance(result, pathlib.Path)
    assert result == out


def test_default_output_path(lulu_pipeline, tmp_path, monkeypatch):
    """When output_path is None, file is saved as <ticker>_DCF.xlsx."""
    stmts, profile, asmp, proj, dcf = lulu_pipeline
    monkeypatch.chdir(tmp_path)
    path = export_workbook(stmts, profile, asmp, proj, dcf, output_path=None)
    assert path.name == f"{_TICKER}_DCF.xlsx"
    assert path.exists()


# ── Sheet count and names ─────────────────────────────────────────────────────

def test_sheet_count(lulu_workbook):
    wb, _ = lulu_workbook
    assert len(wb.sheetnames) == len(EXPECTED_SHEETS)


def test_sheet_names_exact(lulu_workbook):
    wb, _ = lulu_workbook
    assert wb.sheetnames == EXPECTED_SHEETS


def test_sheet_names_in_order(lulu_workbook):
    wb, _ = lulu_workbook
    for expected, actual in zip(EXPECTED_SHEETS, wb.sheetnames):
        assert expected == actual


# ── Summary sheet ─────────────────────────────────────────────────────────────

def test_summary_sheet_has_ticker(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Summary"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any(_TICKER in v for v in all_values)


def test_summary_sheet_has_company_name(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Summary"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("lululemon" in v.lower() for v in all_values)


def test_summary_sheet_has_coverage(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Summary"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Coverage" in v for v in all_values)


def test_summary_sheet_has_quality_issues(lulu_workbook, lulu_pipeline):
    *_, dcf = lulu_pipeline
    wb, _ = lulu_workbook
    ws = wb["Summary"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    # At least one quality issue message should appear verbatim
    assert any(qi.metric in " ".join(all_values) for qi in dcf.quality_issues)


# ── Historical Financials sheet ───────────────────────────────────────────────

def test_historical_sheet_exists(lulu_workbook):
    wb, _ = lulu_workbook
    assert "Historical Financials" in wb.sheetnames


def test_historical_sheet_has_content(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    non_empty = sum(1 for row in ws.iter_rows() for c in row if c.value is not None)
    assert non_empty > 20


def test_historical_sheet_has_revenue_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    col_a = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("Revenue" in v for v in col_a)


def test_historical_sheet_has_income_statement_section(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("INCOME STATEMENT" in v for v in all_values)


def test_historical_sheet_has_balance_sheet_section(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("BALANCE SHEET" in v for v in all_values)


def test_historical_sheet_has_cash_flow_section(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("CASH FLOW" in v for v in all_values)


def test_historical_sheet_year_headers_present(lulu_workbook, lulu_pipeline):
    stmts, *_ = lulu_pipeline
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    all_values = [c.value for row in ws.iter_rows() for c in row]
    for yr in stmts.years:
        assert yr in all_values, f"Year {yr} missing from Historical sheet"


# ── Forecast sheet ────────────────────────────────────────────────────────────

def test_forecast_sheet_has_scenarios(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Forecast"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("BEAR" in v for v in all_values)
    assert any("BASE" in v for v in all_values)
    assert any("BULL" in v for v in all_values)


def test_forecast_sheet_has_revenue_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Forecast"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Revenue" in v for v in all_values)


def test_forecast_sheet_has_fcff_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Forecast"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("FCFF" in v for v in all_values)


def test_forecast_sheet_has_projection_years(lulu_workbook, lulu_pipeline):
    *_, proj, dcf = lulu_pipeline
    wb, _ = lulu_workbook
    ws = wb["Forecast"]
    all_values = [c.value for row in ws.iter_rows() for c in row]
    for fy in proj.base.years:
        assert fy.year in all_values, f"Forecast year {fy.year} missing"


# ── DCF sheet ─────────────────────────────────────────────────────────────────

def test_dcf_sheet_has_enterprise_value_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Enterprise Value" in v for v in all_values)


def test_dcf_sheet_has_equity_value_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Equity Value" in v for v in all_values)


def test_dcf_sheet_has_price_per_share_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Price Per Share" in v for v in all_values)


def test_dcf_sheet_has_scenario_headers(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("BEAR" in v for v in all_values)
    assert any("BASE" in v for v in all_values)
    assert any("BULL" in v for v in all_values)


# ── DCF Multiple sheet ────────────────────────────────────────────────────────

def test_dcf_multiple_sheet_exists(lulu_workbook):
    wb, _ = lulu_workbook
    assert "DCF Multiple" in wb.sheetnames


def test_dcf_multiple_sheet_has_exit_label(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF Multiple"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("Exit" in v or "Multiple" in v for v in all_values)


# ── Sensitivity sheets ────────────────────────────────────────────────────────

def test_sensitivity_ggm_sheet_exists(lulu_workbook):
    wb, _ = lulu_workbook
    assert "Sensitivity - GGM" in wb.sheetnames


def test_sensitivity_exit_sheet_exists(lulu_workbook):
    wb, _ = lulu_workbook
    assert "Sensitivity - Exit" in wb.sheetnames


def test_sensitivity_ggm_has_wacc_values(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - GGM"]
    col_a = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("%" in v for v in col_a), "WACC % values missing from column A"


def test_sensitivity_ggm_has_growth_rate_headers(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - GGM"]
    row2_values = [str(ws.cell(2, c).value or "") for c in range(1, ws.max_column + 1)]
    assert any("%" in v for v in row2_values), "Growth rate % headers missing from row 2"


def test_sensitivity_exit_has_multiple_headers(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - Exit"]
    row2_values = [str(ws.cell(2, c).value or "") for c in range(1, ws.max_column + 1)]
    assert any("x" in v for v in row2_values), "Exit multiple headers missing from row 2"


def test_sensitivity_ggm_has_numeric_prices(lulu_workbook):
    """At least some cells in the GGM sensitivity table should be numeric."""
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - GGM"]
    numeric_count = sum(
        1 for row in ws.iter_rows(min_row=3) for c in row
        if isinstance(c.value, (int, float))
    )
    assert numeric_count > 0


def test_sensitivity_exit_has_numeric_prices(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - Exit"]
    numeric_count = sum(
        1 for row in ws.iter_rows(min_row=3) for c in row
        if isinstance(c.value, (int, float))
    )
    assert numeric_count > 0


# ── Freeze panes ──────────────────────────────────────────────────────────────

def test_historical_sheet_has_freeze_panes(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Historical Financials"]
    assert ws.freeze_panes is not None


def test_dcf_sheet_has_freeze_panes(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["DCF"]
    assert ws.freeze_panes is not None


def test_sensitivity_ggm_has_freeze_panes(lulu_workbook):
    wb, _ = lulu_workbook
    ws = wb["Sensitivity - GGM"]
    assert ws.freeze_panes is not None


# ── Integration: Ford and VZ also produce valid workbooks ─────────────────────

@pytest.mark.parametrize("ticker,cik,name", [
    ("F",  "0000037996", "Ford Motor"),
    ("VZ", "0000732712", "Verizon"),
])
def test_integration_other_companies(ticker, cik, name, tmp_path):
    with open(config.DATA_DIR / "raw" / f"{cik}.json") as f:
        payload = json.load(f)
    facts   = extract_metrics(cik, payload)
    derived = compute_derived_metrics(cik, facts)
    stmts   = build_statements(cik, ticker, facts, derived, max_years=5)
    profile = build_profile(stmts, name)
    asmp    = build_assumptions_from_profile(profile)
    proj    = build_forecast(asmp, profile)
    inputs  = DCFInputs(
        wacc=0.10,
        terminal_growth_rate=0.025,
        exit_ebitda_multiple=8.0,
        net_debt=20_000e6,           # $20B net debt in raw USD
        diluted_shares=1_000_000_000,  # actual share count
    )
    dcf  = build_dcf(proj, inputs)
    path = export_workbook(stmts, profile, asmp, proj, dcf,
                           output_path=tmp_path / f"{ticker}_test.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == EXPECTED_SHEETS
