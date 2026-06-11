import pytest
from openpyxl import load_workbook

from secpull.model import (
    ModelAssumptions,
    avg_growth,
    build_assumptions,
    historical_revenue,
    project_revenue,
    yoy_growth_rates,
)
from secpull.models import FinancialFact


def _rev(fy: int, value: float, end: str, filed: str = "2024-01-01") -> FinancialFact:
    return FinancialFact(
        cik="0001397187",
        metric="revenue",
        tag_used="RevenueFromContractWithCustomerExcludingAssessedTax",
        value=value,
        unit="USD",
        fiscal_year=fy,
        fiscal_period="FY",
        form="10-K",
        end_date=end,
        filed_date=filed,
    )


LULU = [
    _rev(2020, 4.402e9, "2021-01-31"),
    _rev(2021, 6.257e9, "2022-01-30"),
    _rev(2022, 8.111e9, "2023-01-29"),
    _rev(2023, 9.619e9, "2024-01-28"),
]


# ── historical_revenue ──────────────────────────────────────────────────────

def test_historical_revenue_sorted():
    series = historical_revenue(LULU)
    assert series == [
        (2020, 4.402e9),
        (2021, 6.257e9),
        (2022, 8.111e9),
        (2023, 9.619e9),
    ]


def test_historical_revenue_fy_only():
    quarterly = FinancialFact(
        cik="0001397187", metric="revenue", tag_used="x",
        value=1e9, unit="USD", fiscal_year=2023, fiscal_period="Q1",
        form="10-Q", end_date="2023-04-30", filed_date="2023-05-01",
    )
    series = historical_revenue(LULU + [quarterly])
    assert len(series) == 4
    assert all(fy in (2020, 2021, 2022, 2023) for fy, _ in series)


def test_historical_revenue_dedupes_by_latest_filed():
    restatement = _rev(2022, 9.999e9, "2023-01-29", filed="2025-01-01")
    series = historical_revenue(LULU + [restatement])
    fy2022 = dict(series)[2022]
    assert fy2022 == 9.999e9


# ── yoy_growth_rates ────────────────────────────────────────────────────────

def test_yoy_growth_consecutive():
    series = [(2020, 4.402e9), (2021, 6.257e9)]
    rates = yoy_growth_rates(series)
    assert rates[0] == (2020, None)
    assert rates[1][0] == 2021
    assert abs(rates[1][1] - (6.257e9 / 4.402e9 - 1)) < 1e-9


def test_yoy_growth_gap_yields_none():
    series = [(2020, 4.402e9), (2022, 8.111e9)]
    rates = yoy_growth_rates(series)
    assert rates[0] == (2020, None)
    assert rates[1] == (2022, None)


def test_yoy_growth_first_year_always_none():
    series = [(2023, 9.619e9)]
    rates = yoy_growth_rates(series)
    assert rates == [(2023, None)]


# ── avg_growth ──────────────────────────────────────────────────────────────

def test_avg_growth_uses_last_five():
    rates = [
        (2015, None),
        (2016, 0.50),   # excluded — only last 5 non-None
        (2017, 0.10),
        (2018, 0.20),
        (2019, 0.30),
        (2020, 0.40),
        (2021, 0.05),
    ]
    result = avg_growth(rates, n=5)
    expected = (0.10 + 0.20 + 0.30 + 0.40 + 0.05) / 5
    assert abs(result - expected) < 1e-9


def test_avg_growth_fewer_than_five_ok():
    rates = [(2022, None), (2023, 0.186)]
    assert abs(avg_growth(rates, n=5) - 0.186) < 1e-9


def test_insufficient_data_raises():
    rates = yoy_growth_rates([(2023, 9.619e9)])
    with pytest.raises(ValueError):
        avg_growth(rates)


# ── build_assumptions ────────────────────────────────────────────────────────

def test_build_assumptions_offsets():
    a = build_assumptions(0.10, offset=0.05, n_years_used=3)
    assert isinstance(a, ModelAssumptions)
    assert abs(a.base_growth - 0.10) < 1e-9
    assert abs(a.bear_growth - 0.05) < 1e-9
    assert abs(a.bull_growth - 0.15) < 1e-9
    assert a.n_years_used == 3


# ── project_revenue ──────────────────────────────────────────────────────────

def test_project_revenue_compounds():
    a = build_assumptions(0.10, offset=0.05, n_years_used=1)
    proj = project_revenue(1000.0, 2023, a, years=3)
    base = proj["base"]
    assert base[0] == (2024, 1100.0)
    assert abs(base[1][1] - 1210.0) < 0.01
    assert abs(base[2][1] - 1331.0) < 0.01


def test_project_revenue_bear_lt_base_lt_bull():
    a = build_assumptions(0.10, offset=0.05, n_years_used=1)
    proj = project_revenue(1000.0, 2023, a, years=3)
    for i in range(3):
        assert proj["bear"][i][1] < proj["base"][i][1] < proj["bull"][i][1]


def test_project_revenue_year_labels():
    a = build_assumptions(0.10, offset=0.05, n_years_used=1)
    proj = project_revenue(1000.0, 2023, a, years=3)
    for scenario in ("bear", "base", "bull"):
        assert [yr for yr, _ in proj[scenario]] == [2024, 2025, 2026]


# ── export_model_xlsx ────────────────────────────────────────────────────────

def test_export_model_xlsx_roundtrip(tmp_path):
    from secpull.model_export import export_model_xlsx

    rev_series = [(2021, 6.257e9), (2022, 8.111e9), (2023, 9.619e9)]
    a = build_assumptions(0.10, offset=0.05, n_years_used=3)
    proj = project_revenue(rev_series[-1][1], rev_series[-1][0], a, years=3)

    out = tmp_path / "LULU_model.xlsx"
    export_model_xlsx("LULU", rev_series, a, proj, out)

    wb = load_workbook(out)
    ws = wb.active

    # Assumption cells hold numeric values (not formulas)
    assert isinstance(ws["B9"].value, float)   # bear
    assert isinstance(ws["B10"].value, float)  # base
    assert isinstance(ws["B11"].value, float)  # bull

    # Projection cells hold formulas
    assert str(ws["B16"].value).startswith("=")  # bear year 1
    assert str(ws["B17"].value).startswith("=")  # base year 1
    assert str(ws["B18"].value).startswith("=")  # bull year 1

    # Revenue row has values for all 3 historical years
    assert ws["B5"].value is not None
    assert ws["C5"].value is not None
    assert ws["D5"].value is not None
