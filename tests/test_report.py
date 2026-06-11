import pytest
from openpyxl import load_workbook

from secpull.models import FinancialFact
from secpull.report import build_grid, find_missing
from secpull.export import export_xlsx


def _fact(metric, value, fy, fp, end, unit="USD"):
    return FinancialFact(cik="0001397187", metric=metric, tag_used="x",
                         value=value, unit=unit, fiscal_year=fy,
                         fiscal_period=fp, form="10-K", end_date=end,
                         filed_date="2024-03-21")


FACTS = [
    _fact("revenue",     9.619e9, 2023, "FY", "2024-01-28"),
    _fact("revenue",     8.111e9, 2022, "FY", "2023-01-29"),
    _fact("net_income",  1.550e9, 2023, "FY", "2024-01-28"),
    # net_income missing for FY2022 on purpose
    _fact("eps_diluted", 12.20,   2023, "FY", "2024-01-28", unit="USD/shares"),
]


def test_grid_headers_sorted_ascending():
    headers, _ = build_grid(FACTS, periods="FY")
    assert headers == ["Metric", "FY2022 (end 2023-01-29)",
                       "FY2023 (end 2024-01-28)"]


def test_grid_formats_usd_millions():
    _, rows = build_grid(FACTS, periods="FY")
    revenue_row = next(r for r in rows if r[0] == "revenue")
    assert revenue_row[2] == "$9,619M"


def test_grid_formats_eps_two_decimals():
    _, rows = build_grid(FACTS, periods="FY")
    eps_row = next(r for r in rows if r[0] == "eps_diluted")
    assert eps_row[2] == "$12.20"


def test_grid_absent_value_is_na_never_a_number():
    _, rows = build_grid(FACTS, periods="FY")
    ni_row = next(r for r in rows if r[0] == "net_income")
    assert ni_row[1] == "N/A"


def test_find_missing_lists_gaps():
    missing = find_missing(FACTS, periods="FY")
    assert ("net_income", "FY2022 (end 2023-01-29)") in missing
    assert ("eps_diluted", "FY2022 (end 2023-01-29)") in missing


def test_metric_with_no_data_excluded_entirely():
    _, rows = build_grid(FACTS, periods="FY")
    assert not any(r[0] == "cash" for r in rows)


def test_export_xlsx_roundtrip(tmp_path):
    headers, rows = build_grid(FACTS, periods="FY")
    out = tmp_path / "LULU.xlsx"
    export_xlsx("LULU", headers, rows, find_missing(FACTS), out)
    wb = load_workbook(out)
    ws = wb["Financials"]
    assert ws["A1"].value == "Metric"
    assert any(c.value == "Missing Data"
               for row in ws.iter_rows() for c in row)
