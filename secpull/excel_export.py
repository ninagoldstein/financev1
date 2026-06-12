"""
Professional investment-banking style Excel workbook export.

Sheets produced (in order):
  1. Summary               — company overview, coverage stats, quality issues
  2. Historical Financials — IS, BS, CFS with coverage flags
  3. Forecast              — bear/base/bull projections for key line items
  4. DCF                   — Gordon Growth Model valuation outputs
  5. DCF Multiple          — Exit EBITDA Multiple valuation outputs
  6. Sensitivity - GGM     — implied share price grid: WACC × terminal growth rate
  7. Sensitivity - Exit    — implied share price grid: WACC × exit EBITDA multiple

display_unit controls presentation only — internal model always uses raw USD:
  "USD"     (default) — show actual values with comma formatting (#,##0)
  "millions"          — divide by 1,000,000 for display (#,##0.0)

Static values only — no Excel formulas.
"""
from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from secpull.assumptions import ForecastAssumptions
from secpull.dcf import DCFInputs, DCFResult, ScenarioDCF, build_dcf
from secpull.forecast import ProjectedStatements, ScenarioForecast
from secpull.profile import CompanyProfile
from secpull.statements import HistoricalStatements, StatementLine

# ── Colour palette ────────────────────────────────────────────────────────────

_NAVY    = "1F4E79"   # dark navy — main headers
_BLUE    = "2E75B6"   # medium blue — section headers
_LBLUE   = "BDD7EE"   # light blue — sub-header rows / alternate metric groups
_GREY    = "F2F2F2"   # light grey — alternating rows
_YELLOW  = "FFFF99"   # highlight — current WACC/g in sensitivity tables
_WHITE   = "FFFFFF"
_BLACK   = "000000"

# ── Number formats ────────────────────────────────────────────────────────────

_FMT_M   = '#,##0.0'       # $ millions
_FMT_USD = '#,##0'         # raw USD with commas
_FMT_PCT = '0.0%'          # percentage
_FMT_PPX = '0.0x'          # not a real Excel format; used as a label suffix
_FMT_SH  = '$#,##0.00'     # price per share / EPS
_FMT_MX  = '0.0'           # plain decimal (exit multiples on axis)

# ── Sensitivity axis defaults ─────────────────────────────────────────────────

_WACC_RANGE    = [0.075, 0.080, 0.085, 0.090, 0.095, 0.100, 0.105, 0.110, 0.115, 0.120]
_G_RANGE       = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035]
_EXIT_M_RANGE  = [5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]


# ── Low-level style helpers ───────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = _BLACK, size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _bottom_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)


# ── Cell writer ───────────────────────────────────────────────────────────────

def _cell(
    ws,
    row: int,
    col: int,
    value: Any = None,
    fmt: str | None = None,
    bold: bool = False,
    fill_hex: str | None = None,
    font_color: str = _BLACK,
    align: str = "left",
    border: bool = True,
) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, size=10, name="Calibri")
    c.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=False
    )
    if fill_hex:
        c.fill = _fill(fill_hex)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _thin_border()


def _header_row(
    ws,
    row: int,
    start_col: int,
    values: list[Any],
    fill_hex: str = _NAVY,
    font_color: str = _WHITE,
    align: str = "center",
) -> None:
    for i, v in enumerate(values):
        _cell(ws, row, start_col + i, v,
              bold=True, fill_hex=fill_hex, font_color=font_color, align=align)


def _section_label(ws, row: int, col: int, text: str) -> None:
    _cell(ws, row, col, text, bold=True, fill_hex=_BLUE, font_color=_WHITE, align="left")


def _subheader(ws, row: int, col: int, text: str, n_cols: int = 1) -> None:
    _cell(ws, row, col, text, bold=True, fill_hex=_LBLUE, font_color=_BLACK, align="left")
    for c in range(col + 1, col + n_cols):
        _cell(ws, row, c, None, bold=False, fill_hex=_LBLUE, align="left")


# ── Autofit ───────────────────────────────────────────────────────────────────

def _autofit(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in ws.columns:
        best = min_width
        for cell in col:
            val = cell.value
            if val is not None:
                estimated = max(len(str(val)), 8)
                if best < estimated:
                    best = estimated
        ws.column_dimensions[col[0].column_letter].width = min(best + 3, max_width)


# ── Display-unit helpers ──────────────────────────────────────────────────────

def _disp(value: float | None, unit: str) -> float | None:
    """Apply presentation conversion. Internal model always stores raw USD."""
    if value is None:
        return None
    return value / 1_000_000 if unit == "millions" else value


def _dollar_fmt(unit: str) -> str:
    return _FMT_M if unit == "millions" else _FMT_USD


def _unit_suffix(unit: str) -> str:
    """Title suffix added when displaying in millions."""
    return "  ($ millions)" if unit == "millions" else ""


def _unit_header(unit: str) -> str:
    """Sub-header text shown above year columns."""
    return "$ millions" if unit == "millions" else ""


def _stat_line_value(line: StatementLine, yr: int) -> float | None:
    pt = line.values.get(yr)
    return pt.value if pt is not None else None


# ── 1. SUMMARY ────────────────────────────────────────────────────────────────

def _write_summary(
    wb: openpyxl.Workbook,
    profile: CompanyProfile,
    assumptions: ForecastAssumptions,
    projected: ProjectedStatements,
    dcf: DCFResult,
    display_unit: str,
) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    row = 1
    _cell(ws, row, 1, "INVESTMENT ANALYSIS SUMMARY",
          bold=True, fill_hex=_NAVY, font_color=_WHITE, align="left")
    _cell(ws, row, 2, None, fill_hex=_NAVY)
    ws.row_dimensions[row].height = 22
    row += 1

    # Company metadata
    base_rev = _disp(projected.base_revenue, display_unit)
    rev_label = "Base Revenue ($ millions)" if display_unit == "millions" else "Base Revenue"
    rev_value = (f"${base_rev:,.0f}M" if display_unit == "millions"
                 else f"${base_rev:,.0f}")
    meta = [
        ("Company", profile.name),
        ("Ticker", profile.ticker),
        ("CIK", profile.cik),
        ("Historical Period", f"{min(profile.years)}–{max(profile.years)}"),
        ("Forecast Period", f"{projected.base.years[0].year}–{projected.base.years[-1].year}"),
        (rev_label, rev_value),
    ]
    for label, value in meta:
        _cell(ws, row, 1, label, bold=True, fill_hex=_GREY)
        _cell(ws, row, 2, value)
        row += 1

    row += 1
    _subheader(ws, row, 1, "Coverage Statistics", n_cols=2)
    row += 1
    cov_rows = [
        ("Raw Coverage", f"{profile.raw_coverage_pct:.1f}%"),
        ("Adjusted Coverage", f"{profile.adj_coverage_pct:.1f}%"),
        ("Complete", profile.n_complete),
        ("Partial", profile.n_partial),
        ("Derived", profile.n_derived),
        ("Unreliable", profile.n_unreliable),
        ("Stale", profile.n_stale),
        ("Structural Gaps", profile.n_structural_gap),
        ("Absent", profile.n_absent),
    ]
    for label, value in cov_rows:
        _cell(ws, row, 1, label, bold=False, fill_hex=_GREY)
        _cell(ws, row, 2, value, align="right")
        row += 1

    row += 1
    _subheader(ws, row, 1, "Forecast Assumptions", n_cols=2)
    row += 1
    asmp_rows = [
        ("Base Revenue Growth", f"{assumptions.base_revenue_growth:.1%}"),
        ("Bear Revenue Growth", f"{assumptions.bear_revenue_growth:.1%}"),
        ("Bull Revenue Growth", f"{assumptions.bull_revenue_growth:.1%}"),
        ("Gross Margin",
         f"{assumptions.gross_margin:.1%}" if assumptions.gross_margin is not None else "N/A"),
        ("EBIT Margin", f"{assumptions.ebit_margin:.1%}"),
        ("D&A % Revenue", f"{assumptions.da_pct_revenue:.1%}"),
        ("Effective Tax Rate", f"{assumptions.effective_tax_rate:.1%}"),
        ("Capex % Revenue", f"{assumptions.capex_pct_revenue:.1%}"),
        ("NWC % Revenue", f"{assumptions.nwc_pct_revenue:.1%}"),
    ]
    for label, value in asmp_rows:
        _cell(ws, row, 1, label, fill_hex=_GREY)
        _cell(ws, row, 2, value, align="right")
        row += 1

    row += 1
    _subheader(ws, row, 1, f"Quality Issues ({len(dcf.quality_issues)} total)", n_cols=2)
    row += 1
    for qi in dcf.quality_issues:
        sev_color = {"ERROR": "FF0000", "WARNING": "C55A11", "INFO": _BLACK}.get(
            qi.severity, _BLACK
        )
        _cell(ws, row, 1, f"[{qi.severity}] {qi.metric}", bold=qi.severity == "ERROR",
              font_color=sev_color)
        _cell(ws, row, 2, qi.message)
        row += 1

    ws.freeze_panes = "A2"


# ── 2. HISTORICAL FINANCIALS ──────────────────────────────────────────────────

def _write_hist_section(
    ws,
    start_row: int,
    title: str,
    lines: list[tuple[str, StatementLine]],
    years: list[int],
    bold_metrics: set[str],
    fill_metrics: dict[str, str],
    display_unit: str,
) -> int:
    """Write one statement section. Returns next available row."""
    row = start_row
    n_years = len(years)
    dfmt = _dollar_fmt(display_unit)

    # Section header spanning all year columns
    _cell(ws, row, 1, title, bold=True, fill_hex=_BLUE, font_color=_WHITE)
    for c in range(2, 2 + n_years):
        _cell(ws, row, c, None, fill_hex=_BLUE)
    row += 1

    # Year headers
    _cell(ws, row, 1, _unit_header(display_unit), bold=False, fill_hex=_LBLUE)
    for i, yr in enumerate(years):
        _cell(ws, row, 2 + i, yr, bold=True, fill_hex=_LBLUE, align="center")
    row += 1

    # Metric rows
    for label, line in lines:
        is_bold = label in bold_metrics or (line.metric in bold_metrics if line.metric else False)
        fill_hex = fill_metrics.get(label, fill_metrics.get(
            line.metric if line.metric else "", None
        ))
        flag = " [GAP]" if line.is_structural_gap else (" [STALE]" if line.is_stale else "")
        _cell(ws, row, 1, label + flag, bold=is_bold, fill_hex=fill_hex)
        for i, yr in enumerate(years):
            val = _disp(_stat_line_value(line, yr), display_unit)
            _cell(ws, row, 2 + i,
                  val if not (line.is_structural_gap or line.is_stale) else None,
                  fmt=dfmt, bold=is_bold, fill_hex=fill_hex, align="right")
        row += 1

    return row + 1   # blank separator row


def _write_historical(
    wb: openpyxl.Workbook,
    stmts: HistoricalStatements,
    profile: CompanyProfile,
    display_unit: str,
) -> None:
    ws = wb.create_sheet("Historical Financials")
    ws.column_dimensions["A"].width = 32
    years = stmts.years

    row = 1
    _cell(ws, row, 1,
          f"HISTORICAL FINANCIALS — {profile.ticker}{_unit_suffix(display_unit)}",
          bold=True, fill_hex=_NAVY, font_color=_WHITE)
    for c in range(2, 2 + len(years)):
        _cell(ws, row, c, None, fill_hex=_NAVY)
    ws.row_dimensions[row].height = 20
    row += 2

    is_ = stmts.income_statement
    bs  = stmts.balance_sheet
    cfs = stmts.cash_flow_statement

    _BOLD_IS = {"Revenue", "Gross Profit", "EBITDA", "EBIT", "Net Income",
                "revenue", "gross_profit", "ebitda", "operating_income", "net_income"}
    _FILL_IS = {"Revenue": _LBLUE, "EBITDA": _GREY, "Net Income": _GREY}

    is_lines = [
        ("Revenue",           is_.revenue),
        ("Gross Profit",      is_.gross_profit),
        ("EBIT",              is_.ebit),
        ("D&A",               is_.da),
        ("EBITDA",            is_.ebitda),
        ("Interest Expense",  is_.interest_expense),
        ("Income Tax",        is_.income_tax_expense),
        ("Net Income",        is_.net_income),
        ("EPS (Diluted)",     is_.eps_diluted),
    ]
    row = _write_hist_section(ws, row, "INCOME STATEMENT", is_lines,
                              years, _BOLD_IS, _FILL_IS, display_unit)

    _BOLD_BS = {"Total Assets", "Total Equity", "Total Current Assets",
                "Total Current Liabilities", "Total Liabilities"}
    _FILL_BS = {"Total Assets": _GREY, "Total Equity": _GREY,
                "Total Liabilities": _GREY, "Total Current Assets": _LBLUE,
                "Total Current Liabilities": _LBLUE}

    bs_lines = [
        ("Cash & Equivalents",      bs.cash),
        ("Accounts Receivable",     bs.accounts_receivable),
        ("Inventory",               bs.inventory),
        ("Other Current Assets",    bs.other_current_assets),
        ("Total Current Assets",    bs.total_current_assets),
        ("PP&E, Net",               bs.ppe_net),
        ("Goodwill",                bs.goodwill),
        ("Intangibles, Net",        bs.intangibles_net),
        ("Other Non-Current Assets",bs.other_noncurrent_assets),
        ("Total Assets",            bs.total_assets),
        ("Accounts Payable",        bs.accounts_payable),
        ("Accrued Liabilities",     bs.accrued_liabilities),
        ("Current Portion LTD",     bs.current_portion_ltd),
        ("Total Current Liabilities",bs.total_current_liabilities),
        ("Long-Term Debt",          bs.long_term_debt),
        ("Total Liabilities",       bs.total_liabilities),
        ("Retained Earnings",       bs.retained_earnings),
        ("Total Equity",            bs.total_equity),
    ]
    row = _write_hist_section(ws, row, "BALANCE SHEET", bs_lines,
                              years, _BOLD_BS, _FILL_BS, display_unit)

    _BOLD_CFS = {"CFO", "CFI", "CFF", "FCF", "Free Cash Flow",
                 "Cash From Operations", "Cash From Investing", "Cash From Financing"}
    _FILL_CFS = {"Cash From Operations": _GREY, "FCF": _LBLUE,
                 "Free Cash Flow": _LBLUE}

    cfs_lines = [
        ("Net Income",              cfs.net_income),
        ("D&A",                     cfs.da),
        ("Stock-Based Compensation",cfs.stock_based_compensation),
        ("Change in Working Capital",cfs.change_in_nwc),
        ("Cash From Operations",    cfs.cfo),
        ("Capital Expenditures",    cfs.capex),
        ("Acquisitions",            cfs.acquisitions),
        ("Cash From Investing",     cfs.cfi),
        ("Debt Repayment",          cfs.debt_repayment),
        ("Dividends Paid",          cfs.dividends_paid),
        ("Share Repurchases",       cfs.share_repurchases),
        ("Cash From Financing",     cfs.cff),
        ("Free Cash Flow",          cfs.fcf),
    ]
    row = _write_hist_section(ws, row, "CASH FLOW STATEMENT", cfs_lines,
                              years, _BOLD_CFS, _FILL_CFS, display_unit)

    ws.freeze_panes = "B3"
    _autofit(ws)


# ── 3. FORECAST ───────────────────────────────────────────────────────────────

def _write_scenario_block(
    ws,
    start_row: int,
    start_col: int,
    scenario: ScenarioForecast,
    label: str,
    display_unit: str,
) -> None:
    """Write one scenario block (label + year headers + metric rows)."""
    n = len(scenario.years)
    row = start_row
    dfmt = _dollar_fmt(display_unit)

    # Scenario header (merged visually by filling all year columns)
    _cell(ws, row, start_col, label, bold=True,
          fill_hex=_BLUE, font_color=_WHITE, align="center")
    for c in range(start_col + 1, start_col + n):
        _cell(ws, row, c, None, fill_hex=_BLUE)

    # Year headers
    row += 1
    _cell(ws, row, start_col, "Year", bold=True, fill_hex=_LBLUE, align="center")
    for i, fy in enumerate(scenario.years):
        _cell(ws, row, start_col + 1 + i, fy.year, bold=True,
              fill_hex=_LBLUE, align="center")

    # Metric rows
    metrics = [
        ("Revenue",      [_disp(fy.revenue,      display_unit) for fy in scenario.years], dfmt, False),
        ("Gross Profit", [_disp(fy.gross_profit,  display_unit) for fy in scenario.years], dfmt, False),
        ("EBIT",         [_disp(fy.ebit,          display_unit) for fy in scenario.years], dfmt, True),
        ("D&A",          [_disp(fy.da,            display_unit) for fy in scenario.years], dfmt, False),
        ("EBITDA",       [_disp(fy.ebitda,        display_unit) for fy in scenario.years], dfmt, True),
        ("Net Income",   [_disp(fy.net_income,    display_unit) for fy in scenario.years], dfmt, True),
        ("Capex",        [_disp(fy.capex,         display_unit) for fy in scenario.years], dfmt, False),
        ("ΔNWC",         [_disp(fy.delta_nwc,     display_unit) for fy in scenario.years], dfmt, False),
        ("FCFF",         [_disp(fy.fcff,          display_unit) for fy in scenario.years], dfmt, True),
    ]
    for metric_row_idx, (metric_label, vals, fmt, is_bold) in enumerate(metrics):
        r = row + 1 + metric_row_idx
        fill_hex = _GREY if metric_label in ("EBITDA", "EBIT", "Net Income", "FCFF") else None
        _cell(ws, r, start_col, metric_label, bold=is_bold,
              fill_hex=fill_hex, align="left")
        for i, v in enumerate(vals):
            _cell(ws, r, start_col + 1 + i, v, fmt=fmt, bold=is_bold,
                  fill_hex=fill_hex, align="right")


def _write_forecast(
    wb: openpyxl.Workbook,
    projected: ProjectedStatements,
    assumptions: ForecastAssumptions,
    display_unit: str,
) -> None:
    ws = wb.create_sheet("Forecast")

    row = 1
    _cell(ws, row, 1,
          f"PROJECTED FINANCIALS — {projected.ticker}{_unit_suffix(display_unit)}",
          bold=True, fill_hex=_NAVY, font_color=_WHITE)
    n_total_cols = 3 * (len(projected.base.years) + 1) + 1
    for c in range(2, n_total_cols + 2):
        _cell(ws, row, c, None, fill_hex=_NAVY)
    ws.row_dimensions[row].height = 20
    row += 2

    n_yrs = len(projected.base.years)

    for scenario, label, scol in [
        (projected.bear, "BEAR",  1),
        (projected.base, "BASE",  1 + n_yrs + 1 + 1),
        (projected.bull, "BULL",  1 + (n_yrs + 1 + 1) * 2),
    ]:
        _write_scenario_block(ws, row, scol, scenario, label, display_unit)

    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "B3"
    _autofit(ws)


# ── 4. DCF (GGM) ──────────────────────────────────────────────────────────────

def _write_scenario_dcf_col(
    ws,
    start_row: int,
    col: int,
    sdcf: ScenarioDCF,
    mode: str,  # "ggm" or "exit"
    display_unit: str,
) -> None:
    """Write one scenario's DCF values into a column."""
    dfmt = _dollar_fmt(display_unit)
    if mode == "ggm":
        rows = [
            (_disp(sdcf.sum_pv_fcff,           display_unit), dfmt),
            (_disp(sdcf.pv_terminal_value_gg,  display_unit), dfmt),
            (_disp(sdcf.enterprise_value_gg,   display_unit), dfmt),
            (_disp(sdcf.equity_value_gg,       display_unit), dfmt),
            (sdcf.price_per_share_gg,                         _FMT_SH),
        ]
    else:
        rows = [
            (_disp(sdcf.sum_pv_fcff,            display_unit), dfmt),
            (_disp(sdcf.pv_terminal_value_exit, display_unit), dfmt),
            (_disp(sdcf.enterprise_value_exit,  display_unit), dfmt),
            (_disp(sdcf.equity_value_exit,      display_unit), dfmt),
            (sdcf.price_per_share_exit,                        _FMT_SH),
        ]

    bold_rows = {2, 3, 4, 5}   # TV, EV, equity, price
    for i, (val, fmt) in enumerate(rows):
        r = start_row + i
        is_bold = (i + 1) in bold_rows
        fill_hex = _GREY if i in (2, 4) else None
        display = "N/A" if val is None else val
        _cell(ws, r, col, display, fmt=fmt if val is not None else None,
              bold=is_bold, fill_hex=fill_hex, align="right")


def _write_dcf_sheet(
    wb: openpyxl.Workbook,
    dcf: DCFResult,
    mode: str,   # "ggm" or "exit"
    display_unit: str,
) -> None:
    sheet_name = "DCF" if mode == "ggm" else "DCF Multiple"
    tv_label   = "PV Terminal Value — Gordon Growth" if mode == "ggm" \
                 else "PV Terminal Value — Exit Multiple"
    title_suffix = "GORDON GROWTH MODEL" if mode == "ggm" else "EXIT EBITDA MULTIPLE"
    usfx = _unit_suffix(display_unit)
    dfmt = _dollar_fmt(display_unit)

    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 36

    row = 1
    _cell(ws, row, 1,
          f"DCF VALUATION — {dcf.ticker}  |  {title_suffix}{usfx}",
          bold=True, fill_hex=_NAVY, font_color=_WHITE)
    for c in range(2, 6):
        _cell(ws, row, c, None, fill_hex=_NAVY)
    ws.row_dimensions[row].height = 20
    row += 1

    # Inputs summary
    _cell(ws, row, 1, "WACC", bold=True, fill_hex=_GREY)
    _cell(ws, row, 2, dcf.inputs.wacc, fmt=_FMT_PCT, bold=True, fill_hex=_GREY, align="right")
    row += 1
    _cell(ws, row, 1, "Terminal Growth Rate", bold=True, fill_hex=_GREY)
    _cell(ws, row, 2, dcf.inputs.terminal_growth_rate,
          fmt=_FMT_PCT, bold=True, fill_hex=_GREY, align="right")
    row += 1
    if mode == "exit" and dcf.inputs.exit_ebitda_multiple is not None:
        _cell(ws, row, 1, "Exit EBITDA Multiple", bold=True, fill_hex=_GREY)
        _cell(ws, row, 2, f"{dcf.inputs.exit_ebitda_multiple:.1f}x",
              bold=True, fill_hex=_GREY, align="right")
        row += 1
    net_debt_label = "Net Debt ($ millions)" if display_unit == "millions" else "Net Debt"
    _cell(ws, row, 1, net_debt_label, bold=True, fill_hex=_GREY)
    _cell(ws, row, 2, _disp(dcf.inputs.net_debt, display_unit),
          fmt=dfmt, bold=True, fill_hex=_GREY, align="right")
    row += 1
    if dcf.inputs.diluted_shares is not None:
        shares_label = "Diluted Shares (millions)" if display_unit == "millions" else "Diluted Shares"
        _cell(ws, row, 1, shares_label, bold=True, fill_hex=_GREY)
        # shares: divide by 1M in millions mode so "125.0" → 125M shares; raw otherwise
        shares_disp = (dcf.inputs.diluted_shares / 1_000_000
                       if display_unit == "millions"
                       else dcf.inputs.diluted_shares)
        shares_fmt = _FMT_M if display_unit == "millions" else _FMT_USD
        _cell(ws, row, 2, shares_disp,
              fmt=shares_fmt, bold=True, fill_hex=_GREY, align="right")
        row += 1
    row += 1

    # Scenario headers
    _header_row(ws, row, 1, ["", "BEAR", "BASE", "BULL"])
    row += 1

    # Output row labels + values
    def _metric_label(base: str) -> str:
        return f"{base} ($ millions)" if display_unit == "millions" else base

    labels = [
        (_metric_label("Σ PV of FCFF"),    False),
        (_metric_label(tv_label),           False),
        (_metric_label("Enterprise Value"), True),
        (_metric_label("Equity Value"),     True),
        ("Price Per Share",                 True),
    ]
    for i, (label, is_bold) in enumerate(labels):
        fill_hex = _GREY if is_bold else None
        _cell(ws, row + i, 1, label, bold=is_bold, fill_hex=fill_hex)

    for col_idx, sdcf in enumerate([dcf.bear, dcf.base, dcf.bull], start=2):
        _write_scenario_dcf_col(ws, row, col_idx, sdcf, mode, display_unit)

    ws.freeze_panes = "B3"
    _autofit(ws)


# ── 5 & 6. SENSITIVITY TABLES ─────────────────────────────────────────────────

def _sensitivity_price(
    projected: ProjectedStatements,
    base_inputs: DCFInputs,
    wacc: float,
    param: float,
    mode: str,   # "ggm" or "exit"
) -> float | None:
    """Return base-scenario implied share price for one sensitivity cell."""
    try:
        if mode == "ggm":
            new_inputs = DCFInputs(
                wacc=wacc,
                terminal_growth_rate=param,
                exit_ebitda_multiple=None,
                net_debt=base_inputs.net_debt,
                diluted_shares=base_inputs.diluted_shares,
            )
        else:
            new_inputs = DCFInputs(
                wacc=wacc,
                terminal_growth_rate=base_inputs.terminal_growth_rate,
                exit_ebitda_multiple=param,
                net_debt=base_inputs.net_debt,
                diluted_shares=base_inputs.diluted_shares,
            )
        result = build_dcf(projected, new_inputs)
        return (result.base.price_per_share_gg if mode == "ggm"
                else result.base.price_per_share_exit)
    except (ValueError, ZeroDivisionError):
        return None


def _write_sensitivity(
    wb: openpyxl.Workbook,
    projected: ProjectedStatements,
    dcf: DCFResult,
    mode: str,   # "ggm" or "exit"
    display_unit: str,
) -> None:
    sheet_name = "Sensitivity - GGM" if mode == "ggm" else "Sensitivity - Exit"
    col_label  = "Terminal Growth Rate" if mode == "ggm" else "Exit EBITDA Multiple"
    col_params = _G_RANGE if mode == "ggm" else _EXIT_M_RANGE
    col_fmt    = _FMT_PCT if mode == "ggm" else _FMT_MX
    title      = (
        "SENSITIVITY ANALYSIS — Implied Share Price (Base Scenario, Gordon Growth)"
        if mode == "ggm"
        else "SENSITIVITY ANALYSIS — Implied Share Price (Base Scenario, Exit Multiple)"
    )

    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 10

    row = 1
    _cell(ws, row, 1, title, bold=True, fill_hex=_NAVY, font_color=_WHITE)
    for c in range(2, len(col_params) + 3):
        _cell(ws, row, c, None, fill_hex=_NAVY)
    ws.row_dimensions[row].height = 20
    row += 1

    if dcf.inputs.diluted_shares is None:
        _cell(ws, row, 1, "Price per share unavailable: diluted_shares not provided.",
              bold=True, font_color="FF0000")
        return

    # Sub-header
    _cell(ws, row, 1, f"WACC  ↓   {col_label}  →",
          bold=True, fill_hex=_LBLUE, align="left")
    for i, param in enumerate(col_params):
        label = f"{param:.1%}" if mode == "ggm" else f"{param:.0f}x"
        highlight = (
            (mode == "ggm" and abs(param - dcf.inputs.terminal_growth_rate) < 1e-9)
            or (mode == "exit" and dcf.inputs.exit_ebitda_multiple is not None
                and abs(param - dcf.inputs.exit_ebitda_multiple) < 1e-9)
        )
        _cell(ws, row, 2 + i, label, bold=True,
              fill_hex=_YELLOW if highlight else _LBLUE, align="center")
    row += 1

    # Data rows
    for wacc in _WACC_RANGE:
        highlight_row = abs(wacc - dcf.inputs.wacc) < 1e-9
        wacc_label = f"{wacc:.1%}"
        _cell(ws, row, 1, wacc_label, bold=highlight_row,
              fill_hex=_YELLOW if highlight_row else None, align="right")
        for i, param in enumerate(col_params):
            highlight_cell = (
                highlight_row and (
                    (mode == "ggm" and abs(param - dcf.inputs.terminal_growth_rate) < 1e-9)
                    or (mode == "exit" and dcf.inputs.exit_ebitda_multiple is not None
                        and abs(param - dcf.inputs.exit_ebitda_multiple) < 1e-9)
                )
            )
            price = _sensitivity_price(projected, dcf.inputs, wacc, param, mode)
            display = price if price is not None else "N/A"
            _cell(ws, row, 2 + i, display,
                  fmt=_FMT_SH if price is not None else None,
                  fill_hex=_YELLOW if highlight_cell else None,
                  align="right")
        row += 1

    ws.freeze_panes = "B3"
    _autofit(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def export_workbook(
    stmts: HistoricalStatements,
    profile: CompanyProfile,
    assumptions: ForecastAssumptions,
    projected: ProjectedStatements,
    dcf: DCFResult,
    output_path: pathlib.Path | None = None,
    display_unit: str = "USD",
) -> pathlib.Path:
    """Generate and save the full DCF workbook.

    Args:
        stmts:        Historical three-statement model.
        profile:      Five-year CompanyProfile with coverage stats.
        assumptions:  ForecastAssumptions used for the projection.
        projected:    Bear/base/bull ProjectedStatements.
        dcf:          DCFResult with all scenario valuations.
        output_path:  Destination .xlsx path.  Defaults to
                      ``{ticker}_DCF.xlsx`` in the current directory.
        display_unit: Presentation unit for dollar amounts.
                      "USD" (default) — raw values with comma formatting.
                      "millions"      — divide by 1,000,000 for display.
                      The internal model (DCF engine, forecasts, statements,
                      profiles) always stores raw USD regardless of this setting.

    Returns:
        The resolved Path to the saved workbook.
    """
    if display_unit not in ("USD", "millions"):
        raise ValueError(f"display_unit must be 'USD' or 'millions', got {display_unit!r}")
    if output_path is None:
        output_path = pathlib.Path(f"{profile.ticker}_DCF.xlsx")
    output_path = pathlib.Path(output_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    _write_summary(wb, profile, assumptions, projected, dcf, display_unit)
    _write_historical(wb, stmts, profile, display_unit)
    _write_forecast(wb, projected, assumptions, display_unit)
    _write_dcf_sheet(wb, dcf, mode="ggm", display_unit=display_unit)
    _write_dcf_sheet(wb, dcf, mode="exit", display_unit=display_unit)
    _write_sensitivity(wb, projected, dcf, mode="ggm", display_unit=display_unit)
    _write_sensitivity(wb, projected, dcf, mode="exit", display_unit=display_unit)

    wb.save(output_path)
    return output_path
